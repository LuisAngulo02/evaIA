from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Count, Avg, Q
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import io
from datetime import datetime
from authentication.decoradores import student_required, admin_required, teacher_required

@student_required
def student_reports_view(request):
    """Vista para que los estudiantes vean sus calificaciones y reportes"""
    from apps.presentaciones.models import Presentation
    
    # Obtener presentaciones del estudiante
    presentations = Presentation.objects.filter(
        student=request.user,
        status='GRADED'
    ).select_related('assignment', 'assignment__course', 'graded_by')
    
    # Estadísticas del estudiante
    from django.db.models import Max
    stats = {
        'total_presentations': presentations.count(),
        'average_score': presentations.aggregate(avg=Avg('final_score'))['avg'] or 0,
        'best_score': presentations.aggregate(max=Max('final_score'))['max'] or 0,
        'courses_count': presentations.values('assignment__course').distinct().count()
    }
    
    context = {
        'user': request.user,
        'profile': request.user.profile,
        'presentations': presentations,
        'stats': stats,
    }
    return render(request, 'reportes/student_reports.html', context)


@student_required 
def export_student_grades(request):
    """Exporta las calificaciones del estudiante"""
    from apps.presentaciones.models import Presentation
    
    export_format = request.GET.get('format', 'excel')
    
    presentations = Presentation.objects.filter(
        student=request.user,
        status='GRADED'
    ).select_related('assignment', 'assignment__course', 'graded_by').order_by('-graded_at')
    
    if not presentations.exists():
        from django.contrib import messages
        messages.info(request, 'Aún no tienes calificaciones para exportar.')
        return render(request, 'reportes/no_data.html', {'user': request.user})
    
    if export_format == 'pdf':
        return _export_student_pdf(request.user, presentations)
    else:
        return _export_student_excel(request.user, presentations)


def _export_student_excel(student, presentations):
    """Función auxiliar para exportar calificaciones del estudiante a Excel"""
    data = []
    for presentation in presentations:
        data.append({
            'Curso': presentation.assignment.course.name,
            'Código Curso': presentation.assignment.course.code,
            'Asignación': presentation.assignment.title,
            'Título Presentación': presentation.title,
            'Calificación Final': float(presentation.final_score) if presentation.final_score else 0,
            'Puntaje IA': float(presentation.ai_score) if presentation.ai_score else 0,
            'Puntaje Contenido': float(presentation.content_score) if presentation.content_score else 0,
            'Puntaje Fluidez': float(presentation.fluency_score) if presentation.fluency_score else 0,
            'Puntaje Lenguaje Corporal': float(presentation.body_language_score) if presentation.body_language_score else 0,
            'Puntaje Voz': float(presentation.voice_score) if presentation.voice_score else 0,
            'Retroalimentación': presentation.teacher_feedback or '',
            'Calificado por': presentation.graded_by.get_full_name() if presentation.graded_by else '',
            'Fecha Calificación': presentation.graded_at.strftime('%d/%m/%Y %H:%M') if presentation.graded_at else '',
            'Fecha Subida': presentation.uploaded_at_display.strftime('%d/%m/%Y %H:%M') if presentation.uploaded_at_display else ''
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Mis Calificaciones')
        
        workbook = writer.book
        worksheet = writer.sheets['Mis Calificaciones']
        
        # Estilizar header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'mis_calificaciones_{student.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def _export_student_pdf(student, presentations):
    """Función auxiliar para exportar calificaciones del estudiante a PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1
    )
    
    elements = []
    
    # Título
    title = Paragraph(f"Reporte de Calificaciones - {student.get_full_name() or student.username}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Información del estudiante
    student_info = f"""<b>Estudiante:</b> {student.get_full_name() or student.username}<br/>
    <b>Email:</b> {student.email}<br/>
    <b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
    <b>Total presentaciones:</b> {presentations.count()}"""
    
    info_paragraph = Paragraph(student_info, styles['Normal'])
    elements.append(info_paragraph)
    elements.append(Spacer(1, 20))
    
    # Tabla
    data = [['Curso', 'Asignación', 'Calificación', 'IA Score', 'Fecha Calificación']]
    
    for presentation in presentations:
        data.append([
            presentation.assignment.course.code,
            presentation.assignment.title[:25] + '...' if len(presentation.assignment.title) > 25 else presentation.assignment.title,
            f"{presentation.final_score:.1f}" if presentation.final_score else 'N/A',
            f"{presentation.ai_score:.1f}" if presentation.ai_score else 'N/A',
            presentation.graded_at.strftime('%d/%m/%Y') if presentation.graded_at else 'N/A'
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Estadísticas
    if presentations.exists():
        from django.db.models import Max, Min
        elements.append(Spacer(1, 20))
        avg_score = presentations.aggregate(avg=Avg('final_score'))['avg'] or 0
        max_score = presentations.aggregate(max=Max('final_score'))['max'] or 0
        min_score = presentations.aggregate(min=Min('final_score'))['min'] or 0
        
        stats_text = f"""<b>Mis Estadísticas:</b><br/>
        Promedio personal: {avg_score:.2f}<br/>
        Mi mejor calificación: {max_score:.2f}<br/>
        Mi calificación más baja: {min_score:.2f}"""
        
        stats_paragraph = Paragraph(stats_text, styles['Normal'])
        elements.append(stats_paragraph)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f'mis_calificaciones_{student.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@teacher_required
def teacher_reports_view(request):
    """Vista mejorada de rendimiento y analytics para docentes"""
    from apps.presentaciones.models import Course, Assignment, Presentation, Participant
    from django.db.models import Count, Avg, Q, Sum, Max, Min
    from django.utils import timezone
    from datetime import timedelta
    import json
    
    # Obtener cursos del docente
    teacher_courses = Course.objects.filter(teacher=request.user, is_active=True)
    
    # Obtener asignaciones del docente
    teacher_assignments = Assignment.objects.filter(
        course__teacher=request.user, 
        is_active=True
    )
    
    # Obtener presentaciones de los estudiantes del docente
    teacher_presentations = Presentation.objects.filter(
        assignment__course__teacher=request.user
    ).select_related('student', 'assignment', 'assignment__course')
    
    # Estadísticas generales
    total_presentations = teacher_presentations.count()
    graded_presentations = teacher_presentations.filter(status='GRADED').count()
    
    stats = {
        'total_courses': teacher_courses.count(),
        'total_assignments': teacher_assignments.count(),
        'total_presentations': total_presentations,
        'pending_grading': teacher_presentations.filter(
            Q(status='ANALYZED') | Q(status='UPLOADED')
        ).count(),
        'graded_presentations': graded_presentations,
        'average_score': teacher_presentations.filter(
            final_score__isnull=False
        ).aggregate(avg_score=Avg('final_score'))['avg_score'] or 0,
        'ai_average': teacher_presentations.filter(
            ai_score__isnull=False
        ).aggregate(avg_ai=Avg('ai_score'))['avg_ai'] or 0,
        'completion_rate': (graded_presentations / total_presentations * 100) if total_presentations > 0 else 0,
    }
    
    # Distribución de calificaciones
    if graded_presentations > 0:
        grade_distribution = {
            'excelente': teacher_presentations.filter(final_score__gte=90).count(),
            'muy_bueno': teacher_presentations.filter(final_score__gte=80, final_score__lt=90).count(),
            'bueno': teacher_presentations.filter(final_score__gte=70, final_score__lt=80).count(),
            'regular': teacher_presentations.filter(final_score__gte=60, final_score__lt=70).count(),
            'bajo': teacher_presentations.filter(final_score__lt=60).count(),
        }
    else:
        grade_distribution = {
            'excelente': 0, 'muy_bueno': 0, 'bueno': 0, 'regular': 0, 'bajo': 0
        }
    
    # Tendencia de los últimos 30 días
    last_30_days = timezone.now() - timedelta(days=30)
    recent_presentations = teacher_presentations.filter(uploaded_at__gte=last_30_days)
    weekly_trend = []
    for i in range(4):
        week_start = timezone.now() - timedelta(days=(4-i)*7)
        week_end = week_start + timedelta(days=7)
        week_presentations = recent_presentations.filter(
            uploaded_at__gte=week_start,
            uploaded_at__lt=week_end,
            final_score__isnull=False
        )
        weekly_trend.append({
            'week': f'Semana {i+1}',
            'avg_score': week_presentations.aggregate(avg=Avg('final_score'))['avg'] or 0,
            'count': week_presentations.count()
        })
    
    # Estadísticas por curso (mejoradas)
    course_stats = []
    for course in teacher_courses:
        course_presentations = teacher_presentations.filter(assignment__course=course)
        course_graded = course_presentations.filter(final_score__isnull=False)
        
        # Estudiantes únicos
        unique_students = course_presentations.values('student').distinct().count()
        
        course_stats.append({
            'course': course,
            'presentations_count': course_presentations.count(),
            'students_count': unique_students,
            'average_score': course_graded.aggregate(avg=Avg('final_score'))['avg'] or 0,
            'max_score': course_graded.aggregate(max=Max('final_score'))['max'] or 0,
            'min_score': course_graded.aggregate(min=Min('final_score'))['min'] or 0,
            'pending_count': course_presentations.filter(
                Q(status='ANALYZED') | Q(status='UPLOADED')
            ).count(),
            'analyzed_count': course_presentations.filter(status='ANALYZED').count(),
            'graded_count': course_graded.count(),
        })
    
    # Top performers (mejores estudiantes)
    from django.contrib.auth.models import User
    top_students = []
    students = User.objects.filter(
        groups__name='Estudiante',
        presentations__assignment__course__teacher=request.user
    ).distinct()
    
    for student in students[:10]:
        student_presentations = teacher_presentations.filter(
            student=student,
            final_score__isnull=False
        )
        if student_presentations.exists():
            top_students.append({
                'student': student,
                'avg_score': student_presentations.aggregate(avg=Avg('final_score'))['avg'],
                'presentations_count': student_presentations.count(),
            })
    
    top_students = sorted(top_students, key=lambda x: x['avg_score'], reverse=True)[:5]
    
    # Análisis de participación individual
    total_participants = Participant.objects.filter(
        presentation__assignment__course__teacher=request.user
    ).count()
    
    participant_stats = {
        'total_participants': total_participants,
        'avg_participation_time': Participant.objects.filter(
            presentation__assignment__course__teacher=request.user
        ).aggregate(avg=Avg('time_percentage'))['avg'] or 0,
    }
    
    # Presentaciones recientes (últimas 10)
    recent_activity = teacher_presentations.order_by('-uploaded_at')[:10]
    
    # Convertir datos para gráficos (JSON)
    # IMPORTANTE: Convertir Decimal a float para evitar errores de serialización JSON
    context = {
        'user': request.user,
        'stats': {
            'total_courses': stats['total_courses'],
            'total_assignments': stats['total_assignments'],
            'total_presentations': stats['total_presentations'],
            'pending_grading': stats['pending_grading'],
            'graded_presentations': stats['graded_presentations'],
            'average_score': float(stats['average_score']) if stats['average_score'] else 0,
            'ai_average': float(stats['ai_average']) if stats['ai_average'] else 0,
            'completion_rate': float(stats['completion_rate']) if stats['completion_rate'] else 0,
        },
        'grade_distribution': grade_distribution,
        'grade_distribution_json': json.dumps(list(grade_distribution.values())),
        'weekly_trend': weekly_trend,
        'weekly_trend_json': json.dumps([float(w['avg_score']) if w['avg_score'] else 0 for w in weekly_trend]),
        'weekly_labels_json': json.dumps([w['week'] for w in weekly_trend]),
        'course_stats': [
            {
                'course': cs['course'],
                'presentations_count': cs['presentations_count'],
                'students_count': cs['students_count'],
                'average_score': float(cs['average_score']) if cs['average_score'] else 0,
                'max_score': float(cs['max_score']) if cs['max_score'] else 0,
                'min_score': float(cs['min_score']) if cs['min_score'] else 0,
                'pending_count': cs['pending_count'],
                'analyzed_count': cs['analyzed_count'],
                'graded_count': cs['graded_count'],
            }
            for cs in course_stats
        ],
        'top_students': [
            {
                'student': ts['student'],
                'avg_score': float(ts['avg_score']) if ts['avg_score'] else 0,
                'presentations_count': ts['presentations_count'],
            }
            for ts in top_students
        ],
        'participant_stats': {
            'total_participants': participant_stats['total_participants'],
            'avg_participation_time': float(participant_stats['avg_participation_time']) if participant_stats['avg_participation_time'] else 0,
        },
        'recent_activity': recent_activity,
        'teacher_courses': teacher_courses,
        'teacher_assignments': teacher_assignments,
    }
    return render(request, 'reportes/teacher_reports.html', context)

@admin_required
def admin_reports_view(request):
    """Vista para que los administradores vean reportes del sistema"""
    context = {
        'user': request.user,
    }
    return render(request, 'reportes/admin_reports.html', context)


def no_data_view(request):
    """Vista para mostrar cuando no hay datos para exportar"""
    context = {
        'user': request.user,
    }
    return render(request, 'reportes/no_data.html', context)


@teacher_required
def export_grades_excel(request):
    """Exporta las calificaciones a un archivo Excel"""
    from apps.presentaciones.models import Course, Assignment, Presentation
    
    # Obtener datos del docente
    presentations = Presentation.objects.filter(
        assignment__course__teacher=request.user,
        status='GRADED'
    ).select_related(
        'student', 'assignment', 'assignment__course', 'graded_by'
    ).order_by('assignment__course__name', 'assignment__title', 'student__last_name')
    
    # Crear DataFrame
    data = []
    for presentation in presentations:
        data.append({
            'Curso': presentation.assignment.course.name,
            'Código del Curso': presentation.assignment.course.code,
            'Asignación': presentation.assignment.title,
            'Estudiante': presentation.student.get_full_name() or presentation.student.username,
            'Email Estudiante': presentation.student.email,
            'Título Presentación': presentation.title,
            'Fecha Subida': presentation.uploaded_at_display.strftime('%d/%m/%Y %H:%M') if presentation.uploaded_at_display else 'N/A',
            'Calificación Final': float(presentation.final_score) if presentation.final_score else 0,
            'Calificación IA': float(presentation.ai_score) if presentation.ai_score else 0,
            'Puntaje Contenido': float(presentation.content_score) if presentation.content_score else 0,
            'Puntaje Fluidez': float(presentation.fluency_score) if presentation.fluency_score else 0,
            'Puntaje Lenguaje Corporal': float(presentation.body_language_score) if presentation.body_language_score else 0,
            'Puntaje Voz': float(presentation.voice_score) if presentation.voice_score else 0,
            'Retroalimentación Docente': presentation.teacher_feedback or '',
            'Calificado por': presentation.graded_by.get_full_name() if presentation.graded_by else '',
            'Fecha Calificación': presentation.graded_at.strftime('%d/%m/%Y %H:%M') if presentation.graded_at else '',
            'Estado': presentation.get_status_display(),
            'Entregado Tarde': 'Sí' if presentation.is_late else 'No'
        })
    
    if not data:
        from django.contrib import messages
        messages.warning(request, 'No tienes calificaciones disponibles para exportar en este momento.')
        return render(request, 'reportes/no_data.html', {'user': request.user})
    
    df = pd.DataFrame(data)
    
    # Crear archivo Excel en memoria
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Calificaciones')
        
        # Obtener el workbook y worksheet
        workbook = writer.book
        worksheet = writer.sheets['Calificaciones']
        
        # Estilizar el header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Agregar hoja de estadísticas
        stats_data = {
            'Métrica': ['Total Presentaciones', 'Promedio General', 'Mejor Calificación', 'Peor Calificación'],
            'Valor': [
                len(data),
                f"{df['Calificación Final'].mean():.2f}" if len(data) > 0 else '0',
                f"{df['Calificación Final'].max():.2f}" if len(data) > 0 else '0',
                f"{df['Calificación Final'].min():.2f}" if len(data) > 0 else '0'
            ]
        }
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, index=False, sheet_name='Estadísticas')
    
    output.seek(0)
    
    # Crear respuesta HTTP
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'calificaciones_{request.user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@teacher_required
def export_grades_pdf(request):
    """Exporta las calificaciones a un archivo PDF"""
    from apps.presentaciones.models import Course, Assignment, Presentation
    
    # Obtener datos del docente
    presentations = Presentation.objects.filter(
        assignment__course__teacher=request.user,
        status='GRADED'
    ).select_related(
        'student', 'assignment', 'assignment__course', 'graded_by'
    ).order_by('assignment__course__name', 'assignment__title', 'student__last_name')
    
    if not presentations.exists():
        from django.contrib import messages
        messages.warning(request, 'No tienes calificaciones disponibles para exportar en este momento.')
        return render(request, 'reportes/no_data.html', {'user': request.user})
    
    # Crear PDF en memoria
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        alignment=1  # Center
    )
    
    # Elementos del PDF
    elements = []
    
    # Título
    title = Paragraph("Reporte de Calificaciones", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Información del docente
    teacher_info = f"""<b>Docente:</b> {request.user.get_full_name() or request.user.username}<br/>
    <b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
    <b>Total de presentaciones:</b> {presentations.count()}"""
    
    info_paragraph = Paragraph(teacher_info, styles['Normal'])
    elements.append(info_paragraph)
    elements.append(Spacer(1, 20))
    
    # Tabla de datos
    data = [['Curso', 'Asignación', 'Estudiante', 'Calificación', 'IA Score', 'Fecha']]
    
    for presentation in presentations:
        data.append([
            presentation.assignment.course.code,
            presentation.assignment.title[:20] + '...' if len(presentation.assignment.title) > 20 else presentation.assignment.title,
            presentation.student.get_full_name()[:25] + '...' if len(presentation.student.get_full_name() or '') > 25 else (presentation.student.get_full_name() or presentation.student.username),
            f"{presentation.final_score:.1f}" if presentation.final_score else 'N/A',
            f"{presentation.ai_score:.1f}" if presentation.ai_score else 'N/A',
            presentation.graded_at.strftime('%d/%m/%Y') if presentation.graded_at else 'N/A'
        ])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Estadísticas
    if presentations.exists():
        elements.append(Spacer(1, 20))
        from django.db.models import Max, Min
        avg_score = presentations.aggregate(avg=Avg('final_score'))['avg'] or 0
        max_score = presentations.aggregate(max=Max('final_score'))['max'] or 0
        min_score = presentations.aggregate(min=Min('final_score'))['min'] or 0
        
        stats_text = f"""<b>Estadísticas Generales:</b><br/>
        Promedio: {avg_score:.2f}<br/>
        Calificación más alta: {max_score:.2f}<br/>
        Calificación más baja: {min_score:.2f}"""
        
        stats_paragraph = Paragraph(stats_text, styles['Normal'])
        elements.append(stats_paragraph)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f'calificaciones_{request.user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@teacher_required
def export_course_grades(request, course_id):
    """Exporta las calificaciones de un curso específico"""
    from apps.presentaciones.models import Course, Presentation
    
    course = Course.objects.get(id=course_id, teacher=request.user)
    export_format = request.GET.get('format', 'excel')
    
    presentations = Presentation.objects.filter(
        assignment__course=course,
        status='GRADED'
    ).select_related('student', 'assignment', 'graded_by').order_by('student__last_name')
    
    if export_format == 'pdf':
        return _export_course_pdf(course, presentations)
    else:
        return _export_course_excel(course, presentations)


def _export_course_excel(course, presentations):
    """Función auxiliar para exportar curso a Excel"""
    data = []
    for presentation in presentations:
        data.append({
            'Estudiante': presentation.student.get_full_name() or presentation.student.username,
            'Email': presentation.student.email,
            'Asignación': presentation.assignment.title,
            'Título Presentación': presentation.title,
            'Calificación Final': float(presentation.final_score) if presentation.final_score else 0,
            'IA Score': float(presentation.ai_score) if presentation.ai_score else 0,
            'Retroalimentación': presentation.teacher_feedback or '',
            'Fecha Calificación': presentation.graded_at.strftime('%d/%m/%Y %H:%M') if presentation.graded_at else ''
        })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=f'{course.code}')
        
        workbook = writer.book
        worksheet = writer.sheets[f'{course.code}']
        
        # Estilizar
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'{course.code}_calificaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def _export_course_pdf(course, presentations):
    """Función auxiliar para exportar curso a PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1
    )
    
    elements = []
    
    # Título
    title = Paragraph(f"Calificaciones - {course.name} ({course.code})", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Información del curso
    course_info = f"""<b>Docente:</b> {course.teacher.get_full_name()}<br/>
    <b>Fecha:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}<br/>
    <b>Estudiantes evaluados:</b> {presentations.count()}"""
    
    info_paragraph = Paragraph(course_info, styles['Normal'])
    elements.append(info_paragraph)
    elements.append(Spacer(1, 20))
    
    # Tabla
    data = [['Estudiante', 'Asignación', 'Calificación', 'IA Score', 'Fecha']]
    
    for presentation in presentations:
        data.append([
            presentation.student.get_full_name()[:30] + '...' if len(presentation.student.get_full_name() or '') > 30 else (presentation.student.get_full_name() or presentation.student.username),
            presentation.assignment.title[:25] + '...' if len(presentation.assignment.title) > 25 else presentation.assignment.title,
            f"{presentation.final_score:.1f}" if presentation.final_score else 'N/A',
            f"{presentation.ai_score:.1f}" if presentation.ai_score else 'N/A',
            presentation.graded_at.strftime('%d/%m/%Y') if presentation.graded_at else 'N/A'
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f'{course.code}_calificaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response
